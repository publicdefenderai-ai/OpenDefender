import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = "sentencing-commission-audit.xlsx"

PRIORITY_FILLS = {
    "Confirmed":    PatternFill("solid", fgColor="C6EFCE"),
    "Priority 1":   PatternFill("solid", fgColor="FFEB9C"),
    "Priority 2":   PatternFill("solid", fgColor="FCE4D6"),
    "Pattern Jury": PatternFill("solid", fgColor="DDEBF7"),
    "No Commission":PatternFill("solid", fgColor="F2F2F2"),
}

ROWS = [
    ("Florida",        "Confirmed",    "Florida Legislature - Fla. Stat. 921.0022",                           "https://www.flsenate.gov/Laws/Statutes/2024/921.0022"),
    ("Maryland",       "Confirmed",    "MD State Commission on Criminal Sentencing Policy",                    "https://msccsp.org/Files/Guidelines/offensetable.pdf"),
    ("Virginia",       "Confirmed",    "Virginia State Crime Commission VCC",                                  "http://www.scb.virginia.gov/LIDSinformation/vccsearch.cfm"),
    ("Minnesota",      "Priority 1",   "MN Sentencing Guidelines Commission",                                  "https://mn.gov/sentencing-guidelines/"),
    ("Kansas",         "Priority 1",   "Kansas Sentencing Commission",                                         "http://www.sentencing.ks.gov/"),
    ("North Carolina", "Priority 1",   "NC Sentencing and Policy Advisory Commission",                         "https://ncscc.nccourts.org/"),
    ("Oregon",         "Priority 1",   "Oregon Criminal Justice Commission",                                   "https://cjc.oregon.gov/"),
    ("Washington",     "Priority 1",   "WA Sentencing Guidelines Commission",                                  "https://www.sgc.wa.gov/"),
    ("Pennsylvania",   "Priority 2",   "PA Commission on Sentencing",                                          "https://pcs.la.psu.edu/"),
    ("Arkansas",       "Priority 2",   "AR Sentencing Commission",                                             "https://www.asc.arkansas.gov/"),
    ("Michigan",       "Priority 2",   "MI Dept of Corrections - Sentencing Guidelines",                       "https://www.michigan.gov/corrections/"),
    ("Missouri",       "Priority 2",   "MO Sentencing Advisory Commission",                                    "https://www.courts.mo.gov/"),
    ("Alabama",        "Priority 2",   "AL Sentencing Commission",                                             "http://www.alabamasentencing.org/"),
    ("Illinois",       "Pattern Jury", "IL Supreme Court Committee on Criminal Jury Instructions",             "https://www.illinoiscourts.gov/courts/circuit-court/illinois-pattern-jury-instructions-criminal/"),
    ("California",     "Pattern Jury", "Judicial Council of California - CALCRIM",                             "https://www.courts.ca.gov/partners/juryinstructions.htm"),
    ("New York",       "Pattern Jury", "NY Unified Court System - CJI",                                        "https://www.nycourts.gov/judges/cji/2-PenalLaw/index.shtml"),
    ("Washington",     "Pattern Jury", "WA Supreme Court Committee - WPIC",                                    "https://govt.westlaw.com/wccji/Index"),
    ("Delaware",       "Pattern Jury", "Delaware Superior Court",                                               "https://courts.delaware.gov/Superior/pattern/pattern_criminal.aspx"),
    ("North Carolina", "Pattern Jury", "UNC School of Government - NC Pattern Jury Instructions",              "https://www.sog.unc.edu/resources/microsites/north-carolina-pattern-jury-instructions/"),
    ("Texas",          "No Commission","Texas Legislature - Penal Code Title 5",                                "https://statutes.capitol.texas.gov/Docs/PE/htm/PE.19.htm"),
    ("Georgia",        "No Commission","Georgia General Assembly - Criminal Code",                              "https://advance.lexis.com/container?config=00JABmNDIwNjIyYy0yYzc3LTRlMDAtOGJjYi1kMmExMGM5MTE4ODEKAFBvZENhdGFsb2ezl03wraLE&crid="),
]

COLUMNS = [
    ("state",                 15),
    ("priority",              14),
    ("commission_name",       35),
    ("starting_url",          45),
    ("offense_table_exists",  18),
    ("format",                18),
    ("has_offense_name",      18),
    ("has_statute_citation",  18),
    ("has_charge_class",      18),
    ("approx_offense_count",  18),
    ("last_updated",          18),
    ("download_url",          18),
    ("bulk_download_available", 18),
    ("notes",                 40),
]

FINDING_COLS = {h: i + 1 for i, (h, _) in enumerate(COLUMNS)}

VALIDATIONS = {
    "offense_table_exists":    '"Yes,No,Unclear"',
    "format":                  '"Excel,PDF,SearchableDB,StatuteText,None,Other"',
    "has_offense_name":        '"Yes,No"',
    "has_statute_citation":    '"Yes,No"',
    "has_charge_class":        '"Yes,No"',
    "bulk_download_available": '"Yes,No"',
}

wb = Workbook()

# ── Sheet 1: Instructions ────────────────────────────────────────────────────
ws_inst = wb.active
ws_inst.title = "Instructions"

header_text = "Manual Verification Checklist — State Sentencing Commission Offense Tables"
ws_inst["A1"] = header_text
ws_inst["A1"].font = Font(bold=True, size=14)
ws_inst["A2"] = ""

instructions = [
    "Goal: For each state in the Data sheet, visit the listed URL and confirm whether a",
    "public offense table exists, what format it is in, and what fields it includes.",
    "Download or bookmark any confirmed tables and record findings in the Data sheet.",
    "",
    "- offense_table_exists: Yes / No / Unclear",
    "- format: Excel, PDF, SearchableDB, StatuteText, None, or Other",
    "- has_offense_name: Yes / No (does the table include a charge/offense name column?)",
    "- has_statute_citation: Yes / No (does it include a statute section number, e.g. § 921.0022?)",
    "- has_charge_class: Yes / No (does it include felony/misdemeanor class or severity level?)",
    "- approx_offense_count: your best estimate of total offenses listed (integer or blank)",
    "- last_updated: year or date shown on the document (e.g. 2024 or 2024-07)",
    "- download_url: direct URL to the file or page where you accessed the table",
    "- bulk_download_available: Yes / No (can you download the full table as a file, vs. only browse online?)",
    "- notes: anything unusual, caveats, or important observations",
]

for i, line in enumerate(instructions, start=3):
    ws_inst.cell(row=i, column=1, value=line)

ws_inst.column_dimensions["A"].width = 100
ws_inst.sheet_state = "visible"

# ── Sheet 2: Data ────────────────────────────────────────────────────────────
ws_data = wb.create_sheet("Data")

header_font  = Font(bold=True)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
    cell = ws_data.cell(row=1, column=col_idx, value=col_name)
    cell.font  = header_font
    cell.alignment = header_align
    ws_data.column_dimensions[get_column_letter(col_idx)].width = col_width

ws_data.freeze_panes = "A2"
ws_data.row_dimensions[1].height = 30

num_rows = len(ROWS)

for row_idx, (state, priority, commission, url) in enumerate(ROWS, start=2):
    ws_data.cell(row=row_idx, column=1, value=state)
    priority_cell = ws_data.cell(row=row_idx, column=2, value=priority)
    priority_cell.fill = PRIORITY_FILLS.get(priority, PatternFill())
    ws_data.cell(row=row_idx, column=3, value=commission)
    ws_data.cell(row=row_idx, column=4, value=url)

for col_name, formula in VALIDATIONS.items():
    col_idx = FINDING_COLS[col_name]
    col_letter = get_column_letter(col_idx)
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"{col_letter}2:{col_letter}{num_rows + 1}"
    ws_data.add_data_validation(dv)

wb.save(OUTPUT_FILE)
print(f"Saved: {os.path.abspath(OUTPUT_FILE)}")
